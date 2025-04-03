import mimetypes
import os
import pandas as pd
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
import re
import hashlib

image_directory = "./product_images"
os.makedirs(image_directory, exist_ok=True)

csv_directory = "./starquik_links"

options = Options()
options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--disable-gpu")

def create_driver():
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def download_image(image_url, product_name):
    """Download an image from a URL, convert it to .jpg if necessary, and save it locally with a unique name."""
    try:
        response = requests.get(image_url, stream=True)
        response.raise_for_status()

        # Create a safe filename by hashing the image URL and combining it with the product name
        safe_name = re.sub(r"[^\w\-_\. ]", "_", product_name)[:50]
        url_hash = hashlib.md5(image_url.encode('utf-8')).hexdigest()
        file_name = f"{safe_name}_{url_hash[:8]}.jpg" 
        file_path = os.path.join(image_directory, file_name)

        # Convert image to .jpg if needed
        if response.headers["Content-Type"] == "image/webp":
            # Convert .webp to .jpg
            image = PILImage.open(BytesIO(response.content)).convert("RGB")
            image.save(file_path, "JPEG")
        else:
            # If it's not .webp, convert it to .jpg regardless
            image = PILImage.open(BytesIO(response.content)).convert("RGB")
            image.save(file_path, "JPEG")

        return file_path
    except Exception as e:
        print(f"Error downloading image {image_url}: {e}")
        return None

def extract_product_details(driver, product_url):
    try:
        driver.get(product_url)
        print(f"Loaded URL: {product_url}")

        # Wait for the page to fully load and ensure the slider images are present
        WebDriverWait(driver, 45).until(
            EC.presence_of_element_located((By.CLASS_NAME, "slider"))
        )

        # Scroll to the bottom to ensure all images are loaded if it's part of a lazy-loading setup
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "product-slider-image-container"))
        )

        # Extract the page source after the images are loaded
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")

        # Extract product name
        product_name = soup.find(class_="product-name")
        product_name = product_name.text.strip() if product_name else "N/A"

        # Extracting images from both the main slider and the thumbnail slider
        image_elements = soup.select(".product-slider-image-container img, .navSliderItem img")
        image_urls = [img["src"] for img in image_elements if "src" in img.attrs]

        # Ensure unique URLs by converting to a set and then back to a list
        image_urls = list(set(image_urls))

        # Download images
        image_paths = []
        for idx, image_url in enumerate(image_urls, start=1):
            image_file_name = f"{product_name}_{idx}"
            image_path = download_image(image_url, image_file_name)
            if image_path:
                image_paths.append(image_path)

        return {
            "Product Name": product_name,
            "Image Paths": image_paths,
        }

    except Exception as e:
        print(f"Error processing URL {product_url}: {e}")
        return {
            "Product Name": "N/A",
            "Image Paths": [],
        }

def save_to_excel(data, output_file="bever_with_images.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Details"

    ws.column_dimensions["B"].width = 20

    headers = ["Product Name", "Image"]
    ws.append(headers)

    for row_data in data:
        product_name = row_data["Product Name"]
        image_paths = row_data["Image Paths"]

        row = [product_name]
        ws.append(row)

        img_row = ws.max_row
        col = 2

        for image_path in image_paths:
            if image_path and os.path.exists(image_path):
                try:
                    img = Image(image_path)
                    img.width = 100
                    img.height = 100
                    ws.add_image(img, f"{chr(64 + col)}{img_row}")
                except Exception as e:
                    print(f"Error adding image {image_path} to Excel: {e}")
                    ws.cell(row=img_row, column=col).value = "Error adding image"
                col += 1
            else:
                ws.cell(row=img_row, column=col).value = "No Image"
                col += 1

    wb.save(output_file)
    print(f"Excel file saved: {output_file}")

def main():
    driver = create_driver()
    output_data = []

    try:
        for csv_file in os.listdir(csv_directory):
            if csv_file.endswith(".csv"):
                print(f"Processing file: {csv_file}")
                df = pd.read_csv(os.path.join(csv_directory, csv_file))

                for index, row in df.iterrows():
                    product_url = row["Link"]
                    print(f"Fetching details for: {product_url}")
                    product_details = extract_product_details(driver, product_url)
                    output_data.append(product_details)
    finally:
        driver.quit()

    save_to_excel(output_data)

if __name__ == "__main__":
    main()
